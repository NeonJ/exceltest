# -*- encoding: utf-8 -*-
import requests
import json
import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import logging
import re
import time
import curlify
import os


logger = logging.getLogger(__name__)
logger.setLevel(level=logging.INFO)
handler = logging.FileHandler("Automation.log")  # 日志保存在当前执行路径下，可自定义日志名称
handler.setLevel(logging.INFO)
formatter = logging.Formatter('[%(asctime)s] [devices:%(lineno)d] [%(levelname)s] [%(message)s]')
handler.setFormatter(formatter)

# console = logging.StreamHandler()
# console.setLevel(logging.INFO)
# logger.addHandler(handler)
# logger.addHandler(console)


def extraction_get(start_str, end_str, extraction_data):  # 获取参数化字符
    start = extraction_data.find(start_str)
    if start >= 0:
        start += len(start_str)
        end = extraction_data.find(end_str)
        if end >= 0:
            return extraction_data[start:end].strip()


def func_test(api_url):  # 创建调用方法
    if method == "POST":
        if len(header_content_type) == 0 and data != "":
            return requests.post(url=api_url, headers=headers, data=data)
        elif len(header_content_type) != 0:
            return requests.post(url=api_url, json=param, headers=header_content_type)
        else:
            return requests.post(url=api_url, data=data)
    elif method == "GET":
        return requests.get(url=api_url, params=param, headers=header_content_type)
    elif method == "PUT":
        return requests.put(api_url, params=param, headers=header_content_type)
    elif method == "HEAD":
        return requests.head(api_url, params=param, headers=header_content_type)
    elif method == "DELETE":
        return requests.delete(api_url, params=param, headers=header_content_type)


def response_matching(response_value, expect_value):  # 创建断言方法
    if rule == '包含':
        return json.dumps(response_value).count(str(expect_value)) > 0
    elif rule == '不包含':
        return json.dumps(response_value).count(expect_value) == 0
    elif rule == '返回数据大于':
        return len(json.dumps(response_value)) > expect_value
    elif rule == '返回数据小于':
        return len(json.dumps(response_value)) < expect_value
    elif rule == '字符串出现次数':
        expect_str = expect_value.split(';', 2)[0]
        num = expect_value.split(';', 2)[1]
        logger.info("字符串：{0} 出现次数为 {1}".format(expect_str,
                                               json.dumps(response_value).count(expect_str, 1,
                                                                                len(json.dumps(response_value)))))
        return json.dumps(response_value).count(expect_str, 1, len(json.dumps(response_value))) == int(num)
    else:
        print("断言方法错误")
        sys.exit(0)


if __name__ == '__main__':
    logger.info(
        "------------------------------------- Execute TestCases -----------------------------------------------")
    site = {}
    custom_param = {}
    headers = {}
    param = ''
    data = ''
    api_url = ''
    start_time = time.time()
    end_time = time.time()
    response = {}
    api = ''
    passed = 0
    failed = 0
    skipped = 0

    if len(sys.argv) < 2:
        print("Lost Service Name")
        sys.exit(0)
    action = sys.argv[1]
    if action not in ['ami-api-archive-service', 'ami-api-big-data-service', 'ami-api-config-service',
                      'ami-api-gateway-service', 'ami-api-hes-service', 'ami-api-ivy-app-service',
                      'ami-api-ivy-service', 'ami-api-job-service', 'ami-api-mdm-service', 'ami-api-oss-service',
                      'ami-api-register-service', 'ami-api-system-service', 'ami-interface', 'ami-web', 'hes-645-fep',
                      'hes-api', 'hes-core', 'hes-fep', 'kafka', 'map', 'minio', 'postgis', 'redis-cluster',
                      'redis-ha-server', 'spark-master', 'spark-worker', 'zk']:
        print('Service Not in The List')
        sys.exit(0)
    if action == 'ami-api-archive-service':  # 与脚本放在同一目录可直接写excel名称
        wb = load_workbook("ami-api-archive-service.xlsx", data_only=False)
        sheet = wb.sheetnames
        # print(sheet)
    elif action == 'hes-api':
        wb = load_workbook("hes-api.xlsx", data_only=False)
        sheet = wb.sheetnames
        # print(sheet)
    elif action == 'ami-api-big-data-service':
        wb = load_workbook("ami-api-big-data-service.xlsx", data_only=False)
        sheet = wb.sheetnames
        # print(sheet)
    elif action == 'ami-api-config-service':
        wb = load_workbook("ami-api-config-service.xlsx", data_only=False)
        sheet = wb.sheetnames
        # print(sheet)
    elif action == 'ami-api-gateway-service':
        wb = load_workbook("ami-api-gateway-service.xlsx", data_only=False)
        sheet = wb.sheetnames
        # print(sheet)
    elif action == 'ami-api-hes-service':
        wb = load_workbook("ami-api-hes-service.xlsx", data_only=False)
        sheet = wb.sheetnames
        # print(sheet)
    elif action == 'ami-api-ivy-app-service':
        wb = load_workbook("ami-api-ivy-app-service.xlsx", data_only=False)
        sheet = wb.sheetnames
        # print(sheet)
    elif action == 'ami-api-ivy-service':
        wb = load_workbook("ami-api-ivy-service.xlsx", data_only=False)
        sheet = wb.sheetnames
        # print(sheet)
    elif action == 'ami-api-job-service':
        wb = load_workbook("ami-api-job-service.xlsx", data_only=False)
        sheet = wb.sheetnames
        # print(sheet)
    elif action == 'ami-api-mdm-service':
        wb = load_workbook("ami-api-mdm-service.xlsx", data_only=False)
        sheet = wb.sheetnames
        # print(sheet)
    elif action == 'ami-api-register-service':
        wb = load_workbook("ami-api-register-service.xlsx", data_only=False)
        sheet = wb.sheetnames
        # print(sheet)
    elif action == 'ami-api-system-service':
        wb = load_workbook("ami-api-system-service.xlsx", data_only=False)
        sheet = wb.sheetnames
        # print(sheet)
    elif action == 'ami-interface':
        wb = load_workbook("ami-interface.xlsx", data_only=False)
        sheet = wb.sheetnames
        # print(sheet)
    elif action == 'ami-web':
        wb = load_workbook("ami-web.xlsx", data_only=False)
        sheet = wb.sheetnames
        # print(sheet)
    elif action == 'hes-core':
        wb = load_workbook("hes-core.xlsx", data_only=False)
        sheet = wb.sheetnames
        # print(sheet)
    elif action == 'hes-fep':
        wb = load_workbook("hes-fep.xlsx", data_only=False)
        sheet = wb.sheetnames
        # print(sheet)
    elif action == 'hes-core':
        wb = load_workbook("hes-core.xlsx", data_only=False)
        sheet = wb.sheetnames
        # print(sheet)
    else:
        print("Serice No in The List")
        sys.exit(0)

    for row in wb[u"全局参数化配置"].rows:  # excel表格对应的全局参数获取
        site[row[0].value] = row[1].value
    api_base_url = '{api_host}'.format(**site)
    extraction_dict = site
    for i in sheet:
        logger.info(
            "====================================== [API用例表：%s] ===================================================" % (
                i))
        if i == u"测试报告":  # 此字段是excel表格对应表格，拖动至"测试报告总概况"后则不执行
            break

        sheet = wb[i]
        count = 1
        str_checked = "B"
        str_firsttitle = "C"
        str_title = "D"
        str_desc = "E"
        str_method = "F"
        str_API = "G"
        str_headers = "H"
        str_param = "I"
        str_expected_code = "J"
        str_status_code = "K"
        str_error_info = "L"
        str_expect = "M"
        str_rule = "N"
        str_result = "O"
        str_Extraction = "P"
        str_files_name = "Q"
        str_time = "R"
        str_curl = "T"
        warn_fill = PatternFill(fill_type="solid", fgColor="FF0000")
        # verbose = False

        # 命令行参数，-V or --verbose，打印http 的请求结果
        # opts, args = getopt.getopt(sys.argv[1:], "V", ["verbose"])
        # for op, value in opts:
        #     if op == "-V" or op == "--verbose":
        #         verbose = True
        #     else:
        #         sys.exit()

        while count < sheet.max_row:
            str_index = str(count + 1)
            checked = sheet[str_checked + str_index].value

            # 用例是否执行
            if checked == u"No":
                logger.info("[用例编号：%03d] [在excel表格B列设置为不执行，跳过！]" % count)
                count += 1
                skipped += 1
                continue
            try:
                method = sheet[str_method + str_index].value
                rule = sheet[str_rule + str_index].value
                expect = sheet[str_expect + str_index].value
                if sheet[str_method + str_index].value == None:
                    pass
                if "{{" in sheet[str_headers + str_index].value:
                    extraction_param = extraction_get('{{', '}}', sheet[str_headers + str_index].value)
                    header1 = sheet[str_headers + str_index].value.replace('{{', '')
                    header2 = header1.replace('}}', '')
                    header_content_type = json.loads(
                        header2.replace(extraction_param, extraction_dict[extraction_param]))
                elif sheet[str_headers + str_index].value == None:
                    pass
                else:
                    header_content_type = json.loads(sheet[str_headers + str_index].value)

                if "{" in sheet[str_API + str_index].value:
                    api = (sheet[str_API + str_index].value.strip().format(**site).replace(' ', ''))
                elif sheet[str_API + str_index].value == None:
                    pass
                else:
                    api = sheet[str_API + str_index].value.strip()

                # if "{{" in sheet[str_param + str_index].value:
                #     param = json.loads(sheet[str_param + str_index].value.format(**site).replace(' ', ''))
                # elif sheet[str_param + str_index].value == None:
                #     pass
                # # elif "------" in sheet[str_param + str_index].value:
                # #     param = sheet[str_param + str_index].value.replace(' ', '')
                # else:
                #     param = json.loads(sheet[str_param + str_index].value.replace(' ', ''))
                if sheet[str_param + str_index].value == None:
                    param = {}
                elif "{{" in sheet[str_param + str_index].value:
                    extraction_param = extraction_get('{{', '}}', sheet[str_param + str_index].value)
                    param1 = sheet[str_headers + str_index].value.replace('{{', '')
                    param2 = param1.replace('}}', '')
                    param = json.loads(
                        param2.replace(extraction_param, extraction_dict[extraction_param]))
                else:
                    param = json.loads(sheet[str_param + str_index].value)

                expected_code = sheet[str_expected_code + str_index].value

                # if sheet[str_files_name + str_index].value != None:
                #     if "{" in sheet[str_files_name + str_index].value:
                #         file_d = sheet[str_files_name + str_index].value
                #         files = {}
                #         for k, v in file_d.items():
                #             if len(param) == 0:
                #                 files = {k: (file_d[k], open(file_d[k], 'rb'))}
                #             else:
                #                 p_k = param.keys()[0]
                #                 files[p_k] = param[p_k]
                #                 files[k] = (file_d[k], open(file_d[k], 'rb'))
                #
                #         # 生成可用于multipart/form-data上传的数据
                #         data = MultipartEncoder(files)
                #         # 自动生成Content-Type类型和随机码
                #         headers['Content-Type'] = data.content_type
                # else:
                #     pass
                #
                # if "i_" in sheet[str_files_name + str_index].value:
                #     file_d = sheet[str_files_name + str_index].value
                #     files = {}
                #     for k, v in file_d.items():
                #         if len(param) == 0:
                #             files = {k: (file_d[k], open(file_d[k], 'rb'))}
                #         else:
                #             p_k = param.keys()[0]
                #             files[p_k] = param[p_k]
                #             files[k] = (file_d[k], open(file_d[k], 'rb'))
                #     # 生成可用于multipart/form-data上传的数据
                #     data = MultipartEncoder(files)
                # else:
                #     pass

                res = ''
                try:
                    if "http://" in api:
                        res = func_test(api)
                    else:
                        res = func_test(api_base_url + api)
                except Exception as e:
                    logging.error('[接口请求出现异常: %s]' % e)
                    sheet[str_error_info + str_index] = e.__str__()
                    count += 1
                    continue

                # 如果verbose，则打印详细信息
                # if verbose:
                #     print(res.content)
                # sheet[str_status_code + str_index] = res.status_code

                # 返回的状态码和预期的不一样，颜色标红
                if res.status_code != expected_code:
                    # sheet[str_status_code + str_index].fill = warn_fill
                    sheet[str_error_info + str_index].fill = warn_fill
                    sheet[str_error_info + str_index] = res.status_code
                # elif res.content == '':
                # 	sheet[str_status_code + str_index] = res.status_code
                elif sheet[str_files_name + str_index].value == None:
                    sheet[str_status_code + str_index] = res.content
                elif "A" or "B" in sheet[str_files_name + str_index].value:  # 判断该请求是否为下载文件，如果是，res.content为图片
                    file1 = sheet[str_files_name + str_index].value
                    with open(file1, "wb") as f:
                        # f.write(res.content)
                        for chunk in res.iter_content(chunk_size=1024):
                            if chunk:
                                f.write(chunk)
                    sheet[str_status_code + str_index] = str(res.status_code)  # 如果没有返回值，将返回的状态码填入
                else:
                    sheet[str_status_code + str_index] = res.content  # 将返回值放入excel中
                    res.encoding = 'utf-8'
                    ress = res.json()
                sheet[str_time + str_index] = res.elapsed.total_seconds()
                sheet[str_curl + str_index] = curlify.to_curl(res.request)
                res.encoding = 'utf-8'
                ress = res.json()
                response = ress['data']
                sheet[str_result + str_index] = response_matching(ress, expect)

                if response_matching(ress, expect):
                    passed += 1
                else:
                    failed += 1

                logger.info("[用例编号：%03d] [用例标题:%s:%s] >>> %s: %s %s done!" % (
                    count, sheet[str_firsttitle + str_index].value, sheet[str_title + str_index].value, method,
                    api_base_url + api, str(param)))
                logger.info("[用例编号：%03d] [接口响应值] >>> %s" % (count, ress))
                logger.info("[用例编号：%03d] >>> [响应code: %d] [响应时间: %fs]" % (
                    count, res.status_code, res.elapsed.total_seconds()))
                logger.info("-" * 80)
            except Exception as e:
                logger.error("[用例编号：%03d] >>> [请求出错原因：如某个对应excel用例提取失败或对应excel参数json有误：%s]" % (count, e))
                logger.info("-" * 80)

            count += 1
            extraction_value = sheet[str_Extraction + str_index].value
            if extraction_value == None:
                pass
            elif "," in extraction_value:
                extraction_value = re.split(',', extraction_value)
                for extraction in extraction_value:
                    if extraction in response:
                        extraction_dict[extraction] = response[extraction]
                    else:
                        pass
                logger.info("extraction_dict: {}".format(extraction_dict))
            else:
                if extraction_value in response:
                    extraction_dict[extraction_value] = response[extraction_value]
                logger.info("extraction_dict: {}".format(extraction_dict))
            now = time.strftime("%Y-%m-%d", time.localtime(time.time()))
            wb.save(u"TestResult_" + action + "_" + now + ".xlsx")

    logger.info("所有任务结束花费总时间 {:.0f}分 {:.0f}秒".format((time.time() - start_time) // 60, (time.time() - start_time) % 60))

    now = time.strftime("%Y-%m-%d", time.localtime(time.time()))
    # result = load_workbook("TestResult_" + action + "_" + now + ".xlsx", data_only=False, keep_vba=True, read_only=False,
    #                        keep_links=True)
    # sheet = result['测试报告']
    # passed = sheet.cell(4, 4).value
    # failed = sheet.cell(4, 5).value
    # skipped = sheet.cell(4, 6).value
    summary = {'Passed': passed, 'Failed': failed, 'Skipped': skipped,
               'ResultFile': os.path.abspath("TestResult_" + action + "_" + now + ".xlsx")}
    js = json.dumps(summary)
    print(js)